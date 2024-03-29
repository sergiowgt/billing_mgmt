import os
from dataclasses import dataclass
from datetime import datetime
from typing import Any, List
from openpyxl import load_workbook
import pandas as pd
from src.domain.entities.response_error import UtilityBillIgnoredResponse, UtilityBillDuplicatedResponse, UtilityBillOkResponse, UtilityBillErrorResponse
from src.domain.enums import ServiceProviderEnum, ServiceTypeEnum, DocumentTypeEnum
from src.infra.google_drive_handler.Igoogle_drive_handler import IGoogleDriveHandler


@dataclass
class ResultsSaver:
    _log: Any
    _drive: IGoogleDriveHandler

    def __init__(self, log, drive: IGoogleDriveHandler):
        self._log = log
        self._drive = drive

    def _make_google_link(self, file_id: str, file_name: str) -> str:
        #link = f'https://drive.google.com/file/d/{file_id}/view?usp=share_link'
        #return '=HYPERLINK("{}","{}")'.format(link, file_name)
        return f'https://drive.google.com/file/d/{file_id}/view?usp=drive_link'

    def _create_df_qd28(self, list: List[UtilityBillOkResponse]) -> Any:
        def service_type_2_categoria(id_service_type):
            if id_service_type == ServiceTypeEnum.AGUA:
                return 'Água'
            if id_service_type == ServiceTypeEnum.TELECOM:
                return 'Telecomunicações'
            if id_service_type == ServiceTypeEnum.LUZ:
                return 'Eletricidade'
            return ''

        columns = ['#DATA', '#ALOJAMENTO', '#CATEGORIA', '#DESCRICAO', '#VALOR_S/IVA', '#IVA', '#VALOR_C/IVA', '#LINK']
        df = pd.DataFrame(columns=columns)
        now = datetime.now()
        for line in list:
            _dict = {}
            if line.utility_bill.dt_vencimento:
                _dict['#DATA'] = line.utility_bill.dt_vencimento.strftime("%d/%m/%Y")
            else:
                _dict['#DATA'] = line.utility_bill.dt_emissao.strftime("%d/%m/%Y")
            _dict['#ALOJAMENTO'] = line.utility_bill.id_alojamento
            _dict['#CATEGORIA'] = service_type_2_categoria(line.utility_bill.tipo_servico)
            _dict['#DESCRICAO'] = line.utility_bill.periodo_referencia
            _dict['#VALOR_S/IVA'] = ''
            _dict['#IVA'] = ''
            _dict['#VALOR_C/IVA'] = str(line.utility_bill.valor).replace('.', ',')
            _dict['#LINK'] = self._make_google_link(line.google_file_id, line.file_name)
            _dict['Arquivo Original'] = line.file_name
            _dict['Data Processamento'] = now.strftime("%Y/%m/%d.%H:%M:%S")

            df = pd.concat([df, pd.DataFrame.from_records([_dict])])
        return df


    def _create_df_ok(self, list: List[UtilityBillOkResponse]) -> Any:
        columns = ['QQ Destino', 'Alojamento', 'Ano Emissao', 'Mes Emissao', 'Concessionaria', 'Tipo Servico', 'Tipo Documento', 'N. Contrato', 'N. Cliente', 'N. Contribuinte',
                   'Local Consumo', 'Instalacao', 'N. Documento / N. Fatura', 'Periodo Referencia', 'Inicio Referencia', 'Fim Referencia',  'Emissao', 'Vencimento', 'Valor', 'Arquivo Google', 'Arquivo Original', 'Data Processamento']
        df = pd.DataFrame(columns=columns)
        now = datetime.now()
        for line in list:
            _dict = {}
            _dict['QQ Destino'] = 'Sim' if line.utility_bill.is_accounting else 'Não'
            _dict['Alojamento'] = line.utility_bill.id_alojamento
            _dict['Ano Emissao'] = str(line.utility_bill.dt_emissao.year)
            _dict['Mes Emissao'] = format(line.utility_bill.dt_emissao.month, '02d')
            _dict['Concessionaria'] = ServiceProviderEnum(line.utility_bill.concessionaria).name
            _dict['Tipo Servico'] = ServiceTypeEnum(line.utility_bill.tipo_servico).name
            _dict['Tipo Documento'] = DocumentTypeEnum(line.utility_bill.tipo_documento).name
            _dict['N. Contrato'] = line.utility_bill.id_contrato
            _dict['N. Cliente'] = line.utility_bill.id_cliente
            _dict['N. Contribuinte'] = line.utility_bill.id_contribuinte
            _dict['Local Consumo'] = line.utility_bill.local_consumo
            _dict['Instalacao'] = line.utility_bill.instalacao
            _dict['N. Documento / N. Fatura'] = line.utility_bill.id_documento
            _dict['Periodo Referencia'] = line.utility_bill.periodo_referencia
            _dict['Inicio Referencia'] = line.utility_bill.str_inicio_referencia
            _dict['Fim Referencia'] = line.utility_bill.str_fim_referencia
            _dict['Emissao'] = line.utility_bill.str_emissao
            _dict['Vencimento'] = line.utility_bill.str_vencimento
            _dict['Valor'] = line.utility_bill.valor
            _dict['Arquivo Google'] = self._make_google_link(line.google_file_id, line.nome_calculado)
            _dict['Arquivo Original'] = line.file_name
            _dict['Data Processamento'] = now.strftime("%Y/%m/%d.%H:%M:%S")

            df = pd.concat([df, pd.DataFrame.from_records([_dict])])
        return df

    def _create_df_not_found(self, list: List[UtilityBillErrorResponse]) -> Any:
        columns = ['QQ Destino', 'Concessionaria', 'Tipo Servico', 'Tipo Documento', 'N. Contrato', 'N. Cliente', 'N. Contribuinte',
                   'Local Consumo', 'Instalacao', 'N. Documento / N. Fatura', 'Periodo Referencia', 'Inicio Referencia',
                   'Fim Referencia',  'Emissao', 'Vencimento', 'Valor', 'Arquivo Google', 'Arquivo Original']

        df = pd.DataFrame(columns=columns)
        for line in list:
            _dict = {}
            _dict['QQ Destino'] = 'Sim' if line.utility_bill.is_accounting else 'Não'
            _dict['Concessionaria'] = ServiceProviderEnum(line.utility_bill.concessionaria).name
            _dict['Tipo Servico'] = ServiceTypeEnum(line.utility_bill.tipo_servico).name
            _dict['Tipo Documento'] = DocumentTypeEnum(line.utility_bill.tipo_documento).name
            _dict['N. Contrato'] = line.utility_bill.id_contrato
            _dict['N. Cliente'] = line.utility_bill.id_cliente
            _dict['N. Contribuinte'] = line.utility_bill.id_contribuinte
            _dict['Local Consumo'] = line.utility_bill.local_consumo
            _dict['Instalacao'] = line.utility_bill.instalacao
            _dict['N. Documento / N. Fatura'] = line.utility_bill.id_documento
            _dict['Periodo Referencia'] = line.utility_bill.periodo_referencia
            _dict['Inicio Referencia'] = line.utility_bill.str_inicio_referencia
            _dict['Fim Referencia'] = line.utility_bill.str_fim_referencia
            _dict['Emissao'] = line.utility_bill.str_emissao
            _dict['Vencimento'] = line.utility_bill.str_vencimento
            _dict['Valor'] = line.utility_bill.valor
            _dict['Arquivo Google'] = self._make_google_link(line.google_file_id, line.file_name)
            _dict['Arquivo Original'] = line.file_name

            df = pd.concat([df, pd.DataFrame.from_records([_dict])])
        return df

    def _create_df_error(self, list: List[UtilityBillErrorResponse]) -> Any:
        columns = ['Concessionaria', 'Tipo Servico', 'Tipo Documento', 'N. Contrato', 'N. Cliente', 'N. Contribuinte',
                   'Local Consumo', 'Instalacao', 'N. Documento / N. Fatura', 'Periodo Referencia', 'Inicio Referencia',
                   'Fim Referencia',  'Emissao', 'Vencimento', 'Valor', 'Tipo Erro', 'Arquivo Google', 'Arquivo Original']

        df = pd.DataFrame(columns=columns)
        for line in list:
            _dict = {}
            _dict['Concessionaria'] = ServiceProviderEnum(line.utility_bill.concessionaria).name
            _dict['Tipo Servico'] = ServiceTypeEnum(line.utility_bill.tipo_servico).name
            _dict['Tipo Documento'] = DocumentTypeEnum(line.utility_bill.tipo_documento).name
            _dict['N. Contrato'] = line.utility_bill.id_contrato
            _dict['N. Cliente'] = line.utility_bill.id_cliente
            _dict['N. Contribuinte'] = line.utility_bill.id_contribuinte
            _dict['Local Consumo'] = line.utility_bill.local_consumo
            _dict['Instalacao'] = line.utility_bill.instalacao
            _dict['N. Documento / N. Fatura'] = line.utility_bill.id_documento
            _dict['Periodo Referencia'] = line.utility_bill.periodo_referencia
            _dict['Inicio Referencia'] = line.utility_bill.str_inicio_referencia
            _dict['Fim Referencia'] = line.utility_bill.str_fim_referencia
            _dict['Emissao'] = line.utility_bill.str_emissao
            _dict['Vencimento'] = line.utility_bill.str_vencimento
            _dict['Valor'] = line.utility_bill.valor
            _dict['Tipo Erro'] = line.error_type
            _dict['Arquivo Google'] = self._make_google_link(line.google_file_id, line.file_name)
            _dict['Arquivo Original'] = line.file_name

            df = pd.concat([df, pd.DataFrame.from_records([_dict])])
        return df

    def _create_df_duplicated(self, list: List[UtilityBillDuplicatedResponse]) -> Any:
        columns = ['Alojamento', 'Ano Emissao', 'Mes Emissao', 'Concessionaria', 'Tipo Servico', 'Tipo Documento', 'N. Contrato', 'N. Cliente', 'N. Contribuinte',
                   'Local Consumo', 'Instalacao', 'N. Documento / N. Fatura', 'Periodo Referencia', 'Inicio Referencia',
                   'Fim Referencia',  'Emissao', 'Vencimento', 'Valor', 'Tipo', 'Arquivo Google', 'Arquivo Pago', 'Arquivo Original']

        df = pd.DataFrame(columns=columns)
        for line in list:
            _dict = {}
            _dict['Alojamento'] = line.utility_bill.id_alojamento
            _dict['Ano Emissao'] = str(line.utility_bill.dt_emissao.year)
            _dict['Mes Emissao'] = format(line.utility_bill.dt_emissao.month, '02d')
            _dict['Concessionaria'] = ServiceProviderEnum(line.utility_bill.concessionaria).name
            _dict['Tipo Servico'] = ServiceTypeEnum(line.utility_bill.tipo_servico).name
            _dict['Tipo Documento'] = DocumentTypeEnum(line.utility_bill.tipo_documento).name
            _dict['N. Contrato'] = line.utility_bill.id_contrato
            _dict['N. Cliente'] = line.utility_bill.id_cliente
            _dict['N. Contribuinte'] = line.utility_bill.id_contribuinte
            _dict['Local Consumo'] = line.utility_bill.local_consumo
            _dict['Instalacao'] = line.utility_bill.instalacao
            _dict['N. Documento / N. Fatura'] = line.utility_bill.id_documento
            _dict['Periodo Referencia'] = line.utility_bill.periodo_referencia
            _dict['Inicio Referencia'] = line.utility_bill.str_inicio_referencia
            _dict['Fim Referencia'] = line.utility_bill.str_fim_referencia
            _dict['Emissao'] = line.utility_bill.str_emissao
            _dict['Vencimento'] = line.utility_bill.str_vencimento
            _dict['Valor'] = line.utility_bill.valor
            _dict['Tipo'] = line.error_type
            _dict['Arquivo Google'] = self._make_google_link(line.google_file_id, line.file_name)
            _dict['Arquivo Pago'] = line.original_google_link
            _dict['Arquivo Original'] = line.file_name

            df = pd.concat([df, pd.DataFrame.from_records([_dict])])
        return df


    def _create_df_expired(self, list: List[UtilityBillDuplicatedResponse]) -> Any:
        columns = ['Alojamento', 'Ano Emissao', 'Mes Emissao', 'Concessionaria', 'Tipo Servico', 'Tipo Documento', 'N. Contrato', 'N. Cliente', 'N. Contribuinte',
                 'Local Consumo', 'Instalacao', 'N. Documento / N. Fatura', 'Periodo Referencia', 'Inicio Referencia',
                 'Fim Referencia',  'Emissao', 'Vencimento', 'Valor', 'Tipo', 'Arquivo Google', 'Arquivo Original']

        df = pd.DataFrame(columns=columns)
        for line in list:
            _dict = {}
            _dict['Alojamento'] = line.utility_bill.id_alojamento
            _dict['Ano Emissao'] = str(line.utility_bill.dt_emissao.year)
            _dict['Mes Emissao'] = format(line.utility_bill.dt_emissao.month, '02d')
            _dict['Concessionaria'] = ServiceProviderEnum(line.utility_bill.concessionaria).name
            _dict['Tipo Servico'] = ServiceTypeEnum(line.utility_bill.tipo_servico).name
            _dict['Tipo Documento'] = DocumentTypeEnum(line.utility_bill.tipo_documento).name
            _dict['N. Contrato'] = line.utility_bill.id_contrato
            _dict['N. Cliente'] = line.utility_bill.id_cliente
            _dict['N. Contribuinte'] = line.utility_bill.id_contribuinte
            _dict['Local Consumo'] = line.utility_bill.local_consumo
            _dict['Instalacao'] = line.utility_bill.instalacao
            _dict['N. Documento / N. Fatura'] = line.utility_bill.id_documento
            _dict['Periodo Referencia'] = line.utility_bill.periodo_referencia
            _dict['Inicio Referencia'] = line.utility_bill.str_inicio_referencia
            _dict['Fim Referencia'] = line.utility_bill.str_fim_referencia
            _dict['Emissao'] = line.utility_bill.str_emissao
            _dict['Vencimento'] = line.utility_bill.str_vencimento
            _dict['Valor'] = line.utility_bill.valor
            _dict['Tipo'] = line.error_type
            _dict['Arquivo Google'] = self._make_google_link(line.google_file_id, line.file_name)
            _dict['Arquivo Original'] = line.file_name

            df = pd.concat([df, pd.DataFrame.from_records([_dict])])
        return df


    def _create_df_ignored(self, list: List[UtilityBillIgnoredResponse]) -> Any:
        columns = ['Tipo Erro', 'Arquivo Google', 'Arquivo Original']
        df = pd.DataFrame(columns=columns)
        for line in list:
            _dict = {}
            _dict['Tipo Erro'] = line.error_type
            _dict['Arquivo Google'] = self._make_google_link(line.google_file_id, line.file_name)
            _dict['Arquivo Original'] = line.file_name
            df = pd.concat([df, pd.DataFrame.from_records([_dict])])
        return df

    def execute(self, export_filename: str, qd28_filename: str, database_filename: str, new_ok_list: List[UtilityBillOkResponse], not_found_list: List[UtilityBillErrorResponse], error_list: List[UtilityBillErrorResponse],  expired_list: List[UtilityBillErrorResponse], duplicated_list: List[UtilityBillDuplicatedResponse], ignored_list: List[UtilityBillIgnoredResponse], count_contas_pagas: int):
        new_ok_list.sort(key=lambda x: x.utility_bill.concessionaria)
        not_found_list.sort(key=lambda x: x.utility_bill.concessionaria)
        duplicated_list.sort(key=lambda x: x.utility_bill.concessionaria)
        error_list.sort(key=lambda x: x.utility_bill.concessionaria)

        df_ok = self._create_df_ok(new_ok_list)
        df_nf = self._create_df_not_found(not_found_list)
        df_error = self._create_df_error(error_list)
        df_expired = self._create_df_expired(expired_list)
        df_duplicated = self._create_df_duplicated(duplicated_list)
        df_ignored = self._create_df_ignored(ignored_list)
        df_qd28 = self._create_df_qd28(new_ok_list)

        with pd.ExcelWriter(export_filename) as writer:
            df_ok.to_excel(writer, sheet_name='Processados', index=False)
            df_nf.to_excel(writer, sheet_name='Sem Alojamentos', index=False)
            df_error.to_excel(writer, sheet_name='Erros', index=False)
            df_expired.to_excel(writer, sheet_name='Vencidos', index=False)
            df_duplicated.to_excel(writer, sheet_name='Duplicados', index=False)
            df_ignored.to_excel(writer, sheet_name='Ignorados', index=False)

        if os.path.exists(database_filename) is False:
            with pd.ExcelWriter(database_filename) as writer:
                df_ok.to_excel(writer, sheet_name='Database', index=False)
        else:
            if len(new_ok_list) > 0:
                book = load_workbook(database_filename)
                with pd.ExcelWriter(database_filename) as writer:
                    writer.book = book
                    ws = writer.sheets['Database']

                    # adiciona o dataframe na planilha excel
                    for row in df_ok.iterrows():
                        ws.append(row[1].tolist())
                    writer.save()

        if os.path.exists(qd28_filename) is False:
            with pd.ExcelWriter(qd28_filename) as writer:
                df_qd28.to_excel(writer, sheet_name='Página1', index=False)
        else:
            if len(df_qd28) > 0:
                book = load_workbook(qd28_filename)
                with pd.ExcelWriter(qd28_filename) as writer:
                    writer.book = book
                    ws = writer.sheets['Página1']

                    # adiciona o dataframe na planilha excel
                    for row in df_qd28.iterrows():
                        ws.append(row[1].tolist())
                    writer.save()
